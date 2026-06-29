from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import MergedCell

from financial_parser.core.models import Cell, SheetData, Table, TextBlock


def get_merged_cell_value(cell: Cell, merge_map: dict[tuple[int, int], tuple[int, int]]) -> Any:
    """Get the value of a cell, resolving merged cell references."""
    if cell.is_merged and cell.merged_parent:
        return None  # Value handled at parent level
    return cell.value


def build_merge_map(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[tuple[int, int], tuple[int, int]]:
    """Build a mapping from child cells to their merged parent cell."""
    merge_map: dict[tuple[int, int], tuple[int, int]] = {}
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col = merged_range.min_row, merged_range.min_col
        parent = (min_row, min_col)
        for row in range(min_row, merged_range.max_row + 1):
            for col in range(min_col, merged_range.max_col + 1):
                if (row, col) != parent:
                    merge_map[(row, col)] = parent
    return merge_map


def read_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    sheet_name: str,
    resolve_formulas: bool = True,
) -> SheetData:
    """Read a single sheet into SheetData with merged cell handling."""
    merge_map = build_merge_map(ws)
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    merged_ranges = [
        (str(r.min_row) + ":" + str(r.min_col), str(r.max_row) + ":" + str(r.max_col))
        for r in ws.merged_cells.ranges
    ]

    rows: list[list[Cell]] = []
    for row_idx in range(1, max_row + 1):
        row_cells: list[Cell] = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            is_merged = isinstance(cell, MergedCell) or (row_idx, col_idx) in merge_map
            merged_parent = merge_map.get((row_idx, col_idx))

            value = cell.value
            formula = None

            if isinstance(cell, MergedCell):
                if merged_parent:
                    parent_cell = ws.cell(row=merged_parent[0], column=merged_parent[1])
                    value = parent_cell.value
                    formula = parent_cell.formula if hasattr(parent_cell, "formula") else None
            else:
                if hasattr(cell, "formula") and cell.formula:
                    formula = cell.formula
                    if resolve_formulas and cell.value is not None:
                        value = cell.value

            cell_type = "string"
            if isinstance(value, (int, float)):
                cell_type = "number"
            elif isinstance(value, bool):
                cell_type = "boolean"
            elif hasattr(value, "strftime"):
                cell_type = "date"

            row_cells.append(Cell(
                row=row_idx,
                col=col_idx,
                value=value,
                formula=formula,
                data_type=cell_type,
                is_merged=is_merged,
                merged_parent=merged_parent,
            ))
        rows.append(row_cells)

    tables = _extract_tables(rows, sheet_name, ws.title if hasattr(ws, "title") else sheet_name)
    text_blocks = _extract_text_blocks(rows, tables, sheet_name, ws.title if hasattr(ws, "title") else sheet_name)

    return SheetData(
        name=sheet_name,
        merged_cells=merged_ranges,
        total_rows=max_row,
        total_cols=max_col,
        tables=tables,
        text_blocks=text_blocks,
    )


def _extract_tables(rows: list[list[Cell]], sheet_name: str, uri_sheet_name: str) -> list[Table]:
    """Extract logical tables from rows by detecting header rows."""
    if not rows:
        return []

    tables: list[Table] = []
    current_table_start: int | None = None
    headers: list[str] = []

    for idx, row in enumerate(rows):
        non_empty = [c for c in row if c.value is not None and str(c.value).strip()]
        if len(non_empty) >= 2:
            if current_table_start is None:
                current_table_start = idx
                headers = [str(c.value) if c.value else f"Col{i}" for i, c in enumerate(row)]
        else:
            if current_table_start is not None and idx - current_table_start >= 2:
                table_rows = []
                for r in rows[current_table_start:idx]:
                    table_rows.append([c.value for c in r])
                tables.append(Table(
                    sheet_name=sheet_name,
                    start_row=current_table_start + 1,
                    end_row=idx,
                    headers=headers,
                    rows=table_rows,
                    source_uri=f"{uri_sheet_name}!A{current_table_start + 1}:Z{idx}",
                ))
            current_table_start = None
            headers = []

    if current_table_start is not None and len(rows) - current_table_start >= 2:
        table_rows = []
        for r in rows[current_table_start:]:
            table_rows.append([c.value for c in r])
        tables.append(Table(
            sheet_name=sheet_name,
            start_row=current_table_start + 1,
            end_row=len(rows),
            headers=headers,
            rows=table_rows,
            source_uri=f"{uri_sheet_name}!A{current_table_start + 1}:Z{len(rows)}",
        ))

    return tables


def _extract_text_blocks(
    rows: list[list[Cell]],
    tables: list[Table],
    sheet_name: str,
    uri_sheet_name: str,
    min_text_length: int = 10,
) -> list[TextBlock]:
    """Extract text blocks from rows that are not part of tables."""
    if not rows:
        return []

    table_ranges: set[int] = set()
    for table in tables:
        for row_idx in range(table.start_row - 1, table.end_row):
            table_ranges.add(row_idx)

    text_blocks: list[TextBlock] = []
    current_block_start: int | None = None
    current_block_lines: list[str] = []

    for idx, row in enumerate(rows):
        if idx in table_ranges:
            if current_block_start is not None and current_block_lines:
                text_blocks.append(TextBlock(
                    sheet_name=sheet_name,
                    start_row=current_block_start + 1,
                    end_row=idx,
                    content=current_block_lines,
                    source_uri=f"{uri_sheet_name}!A{current_block_start + 1}:Z{idx}",
                ))
            current_block_start = None
            current_block_lines = []
            continue

        text_parts = []
        for cell in row:
            if cell.value is not None:
                val = str(cell.value).strip()
                if val:
                    text_parts.append(val)

        line = " ".join(text_parts)

        if len(line) >= min_text_length:
            if current_block_start is None:
                current_block_start = idx
                current_block_lines = [line]
            else:
                current_block_lines.append(line)
        else:
            if current_block_start is not None and current_block_lines:
                text_blocks.append(TextBlock(
                    sheet_name=sheet_name,
                    start_row=current_block_start + 1,
                    end_row=idx,
                    content=current_block_lines,
                    source_uri=f"{uri_sheet_name}!A{current_block_start + 1}:Z{idx}",
                ))
            current_block_start = None
            current_block_lines = []

    if current_block_start is not None and current_block_lines:
        text_blocks.append(TextBlock(
            sheet_name=sheet_name,
            start_row=current_block_start + 1,
            end_row=len(rows),
            content=current_block_lines,
            source_uri=f"{uri_sheet_name}!A{current_block_start + 1}:Z{len(rows)}",
        ))

    return text_blocks


def read_xlsx(
    file_path: str | Path,
    sheet_names: list[str] | None = None,
    resolve_formulas: bool = True,
) -> tuple[list[SheetData], dict[str, Any]]:
    """Read an XLSX file and return sheet data with metadata."""
    wb = openpyxl.load_workbook(str(file_path), data_only=resolve_formulas)
    metadata: dict[str, Any] = {
        "sheet_names": wb.sheetnames,
        "sheet_count": len(wb.sheetnames),
    }

    sheets: list[SheetData] = []
    target_sheets = sheet_names if sheet_names else wb.sheetnames

    for name in target_sheets:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        sheet_data = read_sheet(ws, name, resolve_formulas)
        sheets.append(sheet_data)

    wb.close()
    return sheets, metadata
